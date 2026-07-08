"""
Vista de Reportes Financieros – Resumen mensual, Diario y por rango
"""
import flet as ft
from datetime import date, datetime, timedelta
from config import PRIMARY, PRIMARY_LT, BG_DARK, BG_CARD, BG_SURFACE, SUCCESS, ERROR, WARNING
from services import api, APIError
from components import loading_icon_button


MONTH_NAMES = [
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
]
MONTH_ABBR = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]


def reports_view(page: ft.Page, app_state: dict):
    cfg      = app_state.get("config", {})
    currency = cfg.get("fiscal.currency_symbol", "$")
    tax_name = cfg.get("fiscal.tax_name", "IVA")

    if not api.is_manager():
        return ft.Container(
            expand=True, bgcolor=BG_DARK, alignment=ft.alignment.center,
            content=ft.Column(horizontal_alignment=ft.CrossAxisAlignment.CENTER, controls=[
                ft.Icon(ft.icons.LOCK, size=64, color=ft.colors.WHITE24),
                ft.Text("Acceso restringido", size=18, color=ft.colors.WHITE54),
            ]),
        )

    daily_content  = ft.Column(expand=True, spacing=10, scroll=ft.ScrollMode.AUTO)
    report_content = ft.Column(expand=True, spacing=10, scroll=ft.ScrollMode.AUTO)
    today_str      = date.today().isoformat()

    # ── Selector de fecha (DatePicker) ────────────────────────────────────────
    # Quitamos los selectores de fecha de "Reportes" agregados en una
    # construcción anterior de esta vista (al navegar entre vistas el overlay
    # de la página no se limpia automáticamente).
    _REPORTS_DP_TAG = "reports_date_picker"
    page.overlay[:] = [c for c in page.overlay if getattr(c, "data", None) != _REPORTS_DP_TAG]
    _date_pickers = []

    def _date_field(value: str, label: str = None, width: int = 135):
        """Campo de fecha de solo lectura que abre un DatePicker al pulsar el ícono de calendario."""
        tf = ft.TextField(
            value=value, label=label, width=width,
            hint_text="AAAA-MM-DD", border_color=PRIMARY, color=ft.colors.WHITE,
            bgcolor=BG_SURFACE, text_size=12, read_only=True,
        )

        try:
            init_date = date.fromisoformat((value or today_str).strip())
        except (ValueError, TypeError):
            init_date = date.today()

        def _on_pick(e):
            if dp.value:
                tf.value = dp.value.date().isoformat() if hasattr(dp.value, "date") else str(dp.value)[:10]
                page.update()

        dp = ft.DatePicker(
            value=init_date,
            first_date=date(2020, 1, 1),
            last_date=date(2100, 12, 31),
            on_change=_on_pick,
            data=_REPORTS_DP_TAG,
        )
        _date_pickers.append(dp)

        def _open(e):
            try:
                dp.value = date.fromisoformat((tf.value or value or today_str).strip())
            except (ValueError, TypeError):
                dp.value = date.today()
            dp.open = True
            page.update()

        tf.suffix = ft.IconButton(
            ft.icons.CALENDAR_MONTH, icon_size=18, icon_color=PRIMARY,
            on_click=_open, tooltip="Elegir fecha",
        )
        return tf, dp

    # Controles de fecha directos (sin ft.Ref) – usados por Reporte Diario / Rango
    daily_date_field, _ = _date_field(today_str, width=135)
    date_from_field, _  = _date_field(today_str, width=135)
    date_to_field, _    = _date_field(today_str, width=135)

    def _show_snack(msg, color=ERROR):
        page.snack_bar = ft.SnackBar(ft.Text(msg, color=ft.colors.WHITE), bgcolor=color, open=True)
        page.update()

    # ── Tarjeta de métricas ───────────────────────────────────────────────────

    def _metric_card(label: str, value: str, icon, color: str, subtitle: str = ""):
        return ft.Container(
            expand=True, bgcolor=BG_CARD,
            border_radius=12,
            border=ft.border.all(1, color + "44"),
            padding=ft.padding.all(18),
            content=ft.Column(spacing=6, controls=[
                ft.Row([
                    ft.Container(
                        width=38, height=38, border_radius=10,
                        bgcolor=color + "22",
                        alignment=ft.alignment.center,
                        content=ft.Icon(icon, color=color, size=20),
                    ),
                    ft.Column(expand=True, spacing=1, controls=[
                        ft.Text(label, size=11, color=ft.colors.WHITE54),
                        ft.Text(value, size=22, color=color, weight=ft.FontWeight.BOLD),
                    ]),
                ], spacing=10),
                ft.Text(subtitle, size=11, color=ft.colors.WHITE38) if subtitle else ft.Container(),
            ]),
        )

    # ── Barra horizontal simple (sin dependencia de chart) ────────────────────

    def _bar_row(label: str, value: float, max_val: float, color: str, is_qty: bool = False):
        pct = (value / max_val) if max_val > 0 else 0
        value_text = f"{value:,.0f}" if is_qty else f"{currency}{value:,.2f}"
        return ft.Container(
            padding=ft.padding.symmetric(vertical=3),
            content=ft.Column(spacing=4, controls=[
                ft.Row([
                    ft.Text(label[:28], size=12, color=ft.colors.WHITE70, expand=True),
                    ft.Text(value_text, size=12, color=color,
                            weight=ft.FontWeight.BOLD, width=90, text_align=ft.TextAlign.RIGHT),
                ]),
                ft.ProgressBar(value=pct, color=color, bgcolor=ft.colors.WHITE12,
                               height=8),
            ]),
        )

    def _empty_placeholder(text: str = "No hay datos para mostrar"):
        return ft.Container(
            alignment=ft.alignment.center, padding=30,
            content=ft.Text(text, color=ft.colors.WHITE38, size=12),
        )

    def _hour_bar(hour: int, count: int, max_count: int):
        pct = (count / max_count) if max_count > 0 else 0
        return ft.Column(
            horizontal_alignment=ft.CrossAxisAlignment.CENTER,
            spacing=2,
            controls=[
                ft.Text(str(count), size=10, color=PRIMARY_LT if count > 0 else ft.colors.WHITE24),
                ft.Container(
                    width=18,
                    height=max(4, int(pct * 80)),
                    bgcolor=PRIMARY if count > 0 else ft.colors.WHITE12,
                    border_radius=ft.border_radius.only(top_left=3, top_right=3),
                ),
                ft.Text(f"{hour:02d}", size=9, color=ft.colors.WHITE38),
            ],
        )

    # ─────────────────────────────────────────────────────────────────────────
    # TAB 0 — Resumen / Dashboard
    # ─────────────────────────────────────────────────────────────────────────

    dash_state = {"year": date.today().year, "monthly": []}
    period_state = {"start": date.today(), "end": date.today()}

    dash_year_text  = ft.Text(str(dash_state["year"]), size=16,
                               color=ft.colors.WHITE, weight=ft.FontWeight.BOLD)
    dash_count_switch = ft.Switch(value=False, active_color=SUCCESS,
                                   tooltip="Mostrar cantidad de ventas en lugar de ingresos")
    dash_chart_container = ft.Container(height=230)
    dash_total_label = ft.Text("Ventas Totales", size=11, color=ft.colors.WHITE54)
    dash_total_text  = ft.Text("0", size=40, weight=ft.FontWeight.BOLD, color=ft.colors.WHITE)
    dash_best_month_text  = ft.Text("---", size=15, color=PRIMARY_LT, weight=ft.FontWeight.BOLD)
    dash_best_month_value = ft.Text(f"{currency}0.00", size=18, color=ft.colors.WHITE,
                                     weight=ft.FontWeight.BOLD)

    period_label_text = ft.Text("", size=13, color=ft.colors.WHITE, weight=ft.FontWeight.BOLD)
    period_top_products   = ft.Column(spacing=6, scroll=ft.ScrollMode.AUTO, height=190)
    period_hours_container = ft.Container(height=140, alignment=ft.alignment.bottom_center)
    period_total_qty_text = ft.Text("0", size=46, weight=ft.FontWeight.BOLD, color=ft.colors.WHITE)
    period_top_categories = ft.Column(spacing=6, scroll=ft.ScrollMode.AUTO, height=160)

    period_from_field, _ = _date_field(today_str, label="Desde", width=150)
    period_to_field, _   = _date_field(today_str, label="Hasta", width=150)

    def _render_dashboard_chart():
        data = dash_state.get("monthly") or [
            {"month": m, "total_sales": 0, "total_revenue": 0.0} for m in range(1, 13)
        ]
        by_count = dash_count_switch.value
        values = [(d["total_sales"] if by_count else d["total_revenue"]) for d in data]
        max_v = max(values) if values and max(values) > 0 else 1

        bar_groups = []
        for i, v in enumerate(values):
            tooltip = (
                f"{MONTH_NAMES[i]}: {v:,.0f}" if by_count
                else f"{MONTH_NAMES[i]}: {currency}{v:,.2f}"
            )
            bar_groups.append(
                ft.BarChartGroup(
                    x=i,
                    bar_rods=[
                        ft.BarChartRod(
                            to_y=v,
                            width=18,
                            color=PRIMARY if v > 0 else ft.colors.WHITE12,
                            border_radius=ft.border_radius.only(top_left=4, top_right=4),
                            tooltip=tooltip,
                        )
                    ],
                )
            )

        dash_chart_container.content = ft.BarChart(
            bar_groups=bar_groups,
            groups_space=10,
            max_y=max_v * 1.2,
            interactive=True,
            expand=True,
            tooltip_bgcolor=ft.colors.with_opacity(0.85, BG_SURFACE),
            border=ft.border.all(0, ft.colors.TRANSPARENT),
            horizontal_grid_lines=ft.ChartGridLines(
                color=ft.colors.WHITE12, width=1, interval=max(max_v / 4, 1)
            ),
            left_axis=ft.ChartAxis(labels_size=44, show_labels=True),
            bottom_axis=ft.ChartAxis(
                labels_size=24,
                labels=[
                    ft.ChartAxisLabel(
                        value=i,
                        label=ft.Text(MONTH_ABBR[i], size=10, color=ft.colors.WHITE54),
                    )
                    for i in range(12)
                ],
            ),
        )

        # Tarjeta "Ventas Totales" + mes de mayor rendimiento
        total_rev   = sum(d["total_revenue"] for d in data)
        total_count = sum(d["total_sales"] for d in data)
        if by_count:
            dash_total_label.value = "Ventas Totales (Cantidad)"
            dash_total_text.value  = f"{total_count:,.0f}"
        else:
            dash_total_label.value = "Ventas Totales"
            dash_total_text.value  = f"{currency}{total_rev:,.2f}"

        best = max(data, key=lambda d: d["total_revenue"], default=None)
        if best and best["total_revenue"] > 0:
            dash_best_month_text.value  = MONTH_NAMES[best["month"] - 1]
            dash_best_month_value.value = f"{currency}{best['total_revenue']:,.2f}"
        else:
            dash_best_month_text.value  = "---"
            dash_best_month_value.value = f"{currency}0.00"

        page.update()

    def load_monthly(e=None):
        try:
            dash_state["monthly"] = api.get_monthly_report(dash_state["year"])
        except APIError as ex:
            dash_state["monthly"] = []
            _show_snack(str(ex))
        _render_dashboard_chart()

    def _year_prev(e=None):
        dash_state["year"] -= 1
        dash_year_text.value = str(dash_state["year"])
        load_monthly()

    def _year_next(e=None):
        dash_state["year"] += 1
        dash_year_text.value = str(dash_state["year"])
        load_monthly()

    def _toggle_dash_mode(e=None):
        _render_dashboard_chart()

    def _period_label():
        d_from = period_state["start"].strftime("%d/%m/%Y")
        d_to   = period_state["end"].strftime("%d/%m/%Y")
        return f"Reportes Periódicos ( {d_from} - {d_to} )"

    def _render_period(data: dict):
        period_label_text.value = _period_label()

        # Top productos
        period_top_products.controls.clear()
        top = data.get("top_products", [])
        if top:
            max_top = max((p.get("total", 0) for p in top), default=0.01)
            for p in top[:8]:
                period_top_products.controls.append(
                    _bar_row(p.get("name", ""), float(p.get("total", 0)), max_top, PRIMARY_LT)
                )
        else:
            period_top_products.controls.append(_empty_placeholder())

        # Ventas por hora
        hours = data.get("sales_by_hour", [])
        max_h = max((h.get("count", 0) for h in hours), default=0)
        if max_h > 0:
            period_hours_container.content = ft.Row(
                controls=[_hour_bar(h["hour"], h["count"], max_h) for h in hours],
                vertical_alignment=ft.CrossAxisAlignment.END,
                spacing=3,
            )
        else:
            period_hours_container.content = _empty_placeholder()

        # Ventas totales (cantidad)
        period_total_qty_text.value = f"{data.get('total_sales', 0):,.0f}"

        # Top grupos de productos (categorías)
        period_top_categories.controls.clear()
        cats = data.get("top_categories", [])
        if cats:
            max_cat = max((c.get("total", 0) for c in cats), default=0.01)
            for c in cats[:10]:
                period_top_categories.controls.append(
                    _bar_row(c.get("name", ""), float(c.get("total", 0)), max_cat, SUCCESS)
                )
        else:
            period_top_categories.controls.append(_empty_placeholder())

        page.update()

    def load_period(e=None):
        try:
            data = api.get_period_report(
                period_state["start"].isoformat(), period_state["end"].isoformat()
            )
            _render_period(data)
        except APIError as ex:
            _show_snack(str(ex))

    def _open_period_picker(e=None):
        period_from_field.value = period_state["start"].isoformat()
        period_to_field.value   = period_state["end"].isoformat()
        err_t = ft.Text("", color=ERROR, size=12)

        def apply_range(_):
            try:
                f = date.fromisoformat((period_from_field.value or "").strip())
                t = date.fromisoformat((period_to_field.value or "").strip())
            except ValueError:
                err_t.value = "Formato de fecha inválido (use AAAA-MM-DD)"
                page.update()
                return
            if f > t:
                err_t.value = "La fecha 'Desde' no puede ser posterior a 'Hasta'"
                page.update()
                return
            period_state["start"] = f
            period_state["end"]   = t
            dlg.open = False
            page.update()
            load_period()

        def quick(days: int):
            period_to_field.value   = date.today().isoformat()
            period_from_field.value = (date.today() - timedelta(days=days - 1)).isoformat()
            page.update()

        dlg = ft.AlertDialog(
            modal=True,
            title=ft.Row([
                ft.Icon(ft.icons.CALENDAR_MONTH, color=PRIMARY),
                ft.Text("Rango de Reportes Periódicos", weight=ft.FontWeight.BOLD),
            ], spacing=8),
            content=ft.Container(
                width=380,
                content=ft.Column(spacing=10, controls=[
                    ft.Row([period_from_field, period_to_field], spacing=10),
                    ft.Row(spacing=4, controls=[
                        ft.TextButton("Hoy", on_click=lambda _: quick(1),
                                      style=ft.ButtonStyle(color=PRIMARY_LT)),
                        ft.TextButton("7 días", on_click=lambda _: quick(7),
                                      style=ft.ButtonStyle(color=PRIMARY_LT)),
                        ft.TextButton("30 días", on_click=lambda _: quick(30),
                                      style=ft.ButtonStyle(color=PRIMARY_LT)),
                        ft.TextButton("Este mes", on_click=lambda _: (
                            setattr(period_from_field, "value", date.today().replace(day=1).isoformat()),
                            setattr(period_to_field, "value", date.today().isoformat()),
                            page.update(),
                        )[-1], style=ft.ButtonStyle(color=PRIMARY_LT)),
                    ]),
                    err_t,
                ]),
            ),
            actions=[
                ft.TextButton("Cancelar",
                              on_click=lambda _: setattr(dlg, "open", False) or page.update()),
                ft.ElevatedButton("Aplicar", icon=ft.icons.CHECK,
                                  on_click=apply_range,
                                  style=ft.ButtonStyle(bgcolor=PRIMARY, color=ft.colors.WHITE)),
            ],
        )
        page.dialog = dlg
        dlg.open = True
        page.update()

    dash_count_switch.on_change = _toggle_dash_mode

    dashboard_content = ft.Column(
        spacing=12,
        controls=[
            ft.Row(spacing=12, controls=[
                # ── Ventas mensuales (gráfica) ──────────────────────────────
                ft.Container(
                    expand=2, height=320, bgcolor=BG_CARD, border_radius=12, padding=16,
                    content=ft.Column(spacing=8, controls=[
                        ft.Row(
                            vertical_alignment=ft.CrossAxisAlignment.CENTER,
                            controls=[
                                ft.Column(expand=True, spacing=2, controls=[
                                    ft.Row(spacing=4, controls=[
                                        ft.Text("Ventas Mensuales -", size=16,
                                                weight=ft.FontWeight.BOLD, color=ft.colors.WHITE),
                                        dash_year_text,
                                    ]),
                                    ft.Text("Datos de ventas agrupados por mes",
                                            size=11, color=ft.colors.WHITE54),
                                ]),
                                loading_icon_button(page, ft.icons.REFRESH, load_monthly,
                                                    icon_color=PRIMARY, tooltip="Actualizar"),
                                dash_count_switch,
                                ft.IconButton(ft.icons.CHEVRON_LEFT, icon_color=ft.colors.WHITE54,
                                              on_click=_year_prev, tooltip="Año anterior"),
                                ft.IconButton(ft.icons.CHEVRON_RIGHT, icon_color=ft.colors.WHITE54,
                                              on_click=_year_next, tooltip="Año siguiente"),
                            ],
                        ),
                        dash_chart_container,
                    ]),
                ),
                # ── Ventas totales / mejor mes ──────────────────────────────
                ft.Container(
                    expand=1, height=320, bgcolor=BG_CARD, border_radius=12, padding=16,
                    content=ft.Column(
                        spacing=10,
                        alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                        controls=[
                            ft.Column(spacing=10, controls=[
                                dash_total_label,
                                dash_total_text,
                            ]),
                            ft.Column(spacing=4, controls=[
                                ft.Divider(color=ft.colors.WHITE12),
                                ft.Text("Mes de mayor rendimiento:", size=11, color=ft.colors.WHITE54),
                                dash_best_month_text,
                                dash_best_month_value,
                            ]),
                        ],
                    ),
                ),
            ]),
            # ── Encabezado de reportes periódicos ───────────────────────────
            ft.Container(
                bgcolor=BG_CARD, border_radius=12,
                padding=ft.padding.symmetric(horizontal=16, vertical=12),
                content=ft.Row(
                    vertical_alignment=ft.CrossAxisAlignment.CENTER,
                    controls=[
                        period_label_text,
                        ft.IconButton(ft.icons.CALENDAR_MONTH, icon_color=PRIMARY,
                                      on_click=_open_period_picker,
                                      tooltip="Cambiar rango de fechas"),
                    ],
                ),
            ),
            # ── Top productos / Ventas por hora / Ventas totales (cantidad) ─
            ft.Row(spacing=12, controls=[
                ft.Container(
                    expand=True, height=290, bgcolor=BG_CARD, border_radius=12, padding=16,
                    content=ft.Column(spacing=8, controls=[
                        ft.Text("Top Productos", size=14, weight=ft.FontWeight.BOLD, color=ft.colors.WHITE),
                        period_top_products,
                    ]),
                ),
                ft.Container(
                    expand=True, height=290, bgcolor=BG_CARD, border_radius=12, padding=16,
                    content=ft.Column(spacing=8, controls=[
                        ft.Text("Ventas por hora", size=14, weight=ft.FontWeight.BOLD, color=ft.colors.WHITE),
                        period_hours_container,
                    ]),
                ),
                ft.Container(
                    expand=True, height=290, bgcolor=BG_CARD, border_radius=12, padding=16,
                    content=ft.Column(
                        horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                        alignment=ft.MainAxisAlignment.CENTER,
                        spacing=8, controls=[
                            ft.Text("Ventas Totales (Cantidad)", size=14,
                                    weight=ft.FontWeight.BOLD, color=ft.colors.WHITE),
                            period_total_qty_text,
                        ],
                    ),
                ),
            ]),
            # ── Top grupos de productos ──────────────────────────────────────
            ft.Container(
                bgcolor=BG_CARD, border_radius=12, padding=16,
                content=ft.Column(spacing=4, controls=[
                    ft.Text("Top grupos de productos", size=14, weight=ft.FontWeight.BOLD, color=ft.colors.WHITE),
                    ft.Text("Grupos de productos de mayor venta en el período seleccionado",
                            size=11, color=ft.colors.WHITE54),
                    ft.Container(height=8),
                    period_top_categories,
                ]),
            ),
        ],
    )

    # ─────────────────────────────────────────────────────────────────────────
    # TAB 1 — Reporte Diario / TAB 2 — Reporte por Rango
    # ─────────────────────────────────────────────────────────────────────────

    def render_daily(data: dict):
        daily_content.controls.clear()
        total_rev  = float(data.get("total_revenue", 0))
        total_tax  = float(data.get("total_tax", 0))
        total_disc = float(data.get("total_discounts", 0))
        total_s    = int(data.get("total_sales", 0))
        cancelled  = int(data.get("cancelled_sales", 0))
        cash_s     = float(data.get("cash_sales", 0))
        card_s     = float(data.get("card_sales", 0))
        trans_s    = float(data.get("transfer_sales", 0))
        avg_ticket = total_rev / total_s if total_s > 0 else 0

        # Tarjetas de métricas
        daily_content.controls.append(ft.Row(spacing=10, controls=[
            _metric_card("Ingresos totales", f"{currency}{total_rev:,.2f}",
                         ft.icons.ATTACH_MONEY, SUCCESS),
            _metric_card("Ventas completadas", str(total_s),
                         ft.icons.RECEIPT_LONG, PRIMARY_LT,
                         f"{cancelled} cancelada(s)"),
            _metric_card("Ticket promedio", f"{currency}{avg_ticket:,.2f}",
                         ft.icons.SHOW_CHART, WARNING),
            _metric_card(tax_name + " generado", f"{currency}{total_tax:,.2f}",
                         ft.icons.RECEIPT, ft.colors.PURPLE_300),
        ]))

        # Desglose por método de pago
        max_pay = max(cash_s, card_s, trans_s, 0.01)
        daily_content.controls.append(ft.Container(
            bgcolor=BG_CARD, border_radius=12, padding=20,
            content=ft.Column(spacing=10, controls=[
                ft.Text("Ventas por método de pago", size=14, color=ft.colors.WHITE,
                        weight=ft.FontWeight.BOLD),
                _bar_row("💵 Efectivo",     cash_s,  max_pay, SUCCESS),
                _bar_row("💳 Tarjeta",      card_s,  max_pay, PRIMARY_LT),
                _bar_row("🏦 Transferencia", trans_s, max_pay, WARNING),
                ft.Row([
                    ft.Text(f"Descuentos aplicados: {currency}{total_disc:,.2f}",
                            size=12, color=ERROR),
                ]),
            ]),
        ))

        # Top productos
        top = data.get("top_products", [])
        if top:
            max_top = max((p.get("total",0) for p in top), default=0.01)
            prod_bars = [
                _bar_row(p.get("name","")[:30], float(p.get("total",0)), max_top, PRIMARY_LT)
                for p in top[:10]
            ]
            daily_content.controls.append(ft.Container(
                bgcolor=BG_CARD, border_radius=12, padding=20,
                content=ft.Column(spacing=8, controls=[
                    ft.Text("Top 10 productos más vendidos", size=14,
                            color=ft.colors.WHITE, weight=ft.FontWeight.BOLD),
                    *prod_bars,
                ]),
            ))

        # Ventas por hora
        hours_data = data.get("sales_by_hour", [])
        if hours_data:
            max_h = max((h.get("count",0) for h in hours_data), default=1)
            hour_bars = [
                _hour_bar(h.get("hour",0), h.get("count",0), max_h)
                for h in hours_data
            ]
            daily_content.controls.append(ft.Container(
                bgcolor=BG_CARD, border_radius=12, padding=20,
                content=ft.Column(spacing=8, controls=[
                    ft.Text("Distribución de ventas por hora", size=14,
                            color=ft.colors.WHITE, weight=ft.FontWeight.BOLD),
                    ft.Container(
                        content=ft.Row(
                            controls=hour_bars,
                            vertical_alignment=ft.CrossAxisAlignment.END,
                            spacing=3,
                        ),
                        height=120,
                        alignment=ft.alignment.bottom_center,
                    ),
                ]),
            ))

        page.update()

    def render_range(data: list):
        report_content.controls.clear()
        if not data:
            report_content.controls.append(ft.Text("Sin datos", color=ft.colors.WHITE54))
            page.update(); return

        total_rev    = sum(float(d.get("total_revenue",0)) for d in data)
        total_s      = sum(int(d.get("total_sales",0)) for d in data)
        total_disc   = sum(float(d.get("total_discounts",0)) for d in data)
        total_tax    = sum(float(d.get("total_tax",0)) for d in data)

        report_content.controls.append(ft.Row(spacing=10, controls=[
            _metric_card("Ingresos del período", f"{currency}{total_rev:,.2f}",
                         ft.icons.TRENDING_UP, SUCCESS),
            _metric_card("Total ventas", str(total_s), ft.icons.RECEIPT_LONG, PRIMARY_LT),
            _metric_card(tax_name, f"{currency}{total_tax:,.2f}", ft.icons.RECEIPT, ft.colors.PURPLE_300),
            _metric_card("Descuentos", f"{currency}{total_disc:,.2f}", ft.icons.DISCOUNT, WARNING),
        ]))

        # Tabla diaria
        rows = []
        for d in data:
            rev = float(d.get("total_revenue",0))
            s   = int(d.get("total_sales",0))
            rows.append(ft.DataRow(cells=[
                ft.DataCell(ft.Text(d.get("date",""), size=12, color=ft.colors.WHITE70)),
                ft.DataCell(ft.Text(str(s), size=13, color=ft.colors.WHITE)),
                ft.DataCell(ft.Text(f"{currency}{rev:,.2f}", size=13, color=SUCCESS,
                                    weight=ft.FontWeight.BOLD)),
                ft.DataCell(ft.Text(f"{currency}{float(d.get('total_tax',0)):,.2f}",
                                    size=12, color=ft.colors.WHITE54)),
                ft.DataCell(ft.Text(f"{currency}{float(d.get('total_discounts',0)):,.2f}",
                                    size=12, color=WARNING)),
                ft.DataCell(ft.Text(f"{currency}{rev/s:.2f}" if s > 0 else "—",
                                    size=12, color=ft.colors.WHITE60)),
            ]))

        report_content.controls.append(ft.Container(
            bgcolor=BG_CARD, border_radius=12, padding=20,
            content=ft.Column(spacing=8, controls=[
                ft.Text("Detalle por día", size=14, color=ft.colors.WHITE, weight=ft.FontWeight.BOLD),
                ft.DataTable(
                    border=ft.border.all(1, ft.colors.WHITE12),
                    border_radius=8,
                    heading_row_color=BG_SURFACE,
                    columns=[
                        ft.DataColumn(ft.Text("Fecha",      color=ft.colors.WHITE70, size=12)),
                        ft.DataColumn(ft.Text("Ventas",     color=ft.colors.WHITE70, size=12), numeric=True),
                        ft.DataColumn(ft.Text("Ingresos",   color=ft.colors.WHITE70, size=12), numeric=True),
                        ft.DataColumn(ft.Text(tax_name,     color=ft.colors.WHITE70, size=12), numeric=True),
                        ft.DataColumn(ft.Text("Descuentos", color=ft.colors.WHITE70, size=12), numeric=True),
                        ft.DataColumn(ft.Text("Ticket prom",color=ft.colors.WHITE70, size=12), numeric=True),
                    ],
                    rows=rows,
                ),
            ]),
        ))
        page.update()

    # ─────────────────────────────────────────────────────────────────────────
    # TAB 3 — Informes (generador de informes personalizados)
    # ─────────────────────────────────────────────────────────────────────────

    PAYMENT_OPTIONS = [
        ("", "Todas"), ("cash", "Efectivo"), ("card", "Tarjeta"),
        ("transfer", "Transferencia"), ("mixed", "Mixto"),
    ]
    PAYMENT_LABELS = dict(PAYMENT_OPTIONS)

    REPORT_GROUPS = [
        ("Ventas", [
            {"key": "productos", "label": "Productos", "filters": ["range", "cashier", "payment_method"]},
            {"key": "grupos_productos", "label": "Grupo de Productos", "filters": ["range", "cashier", "payment_method"]},
            {"key": "cajeros", "label": "Cajeros", "filters": ["range", "payment_method"]},
            {"key": "formas_pago", "label": "Formas de Pagos", "filters": ["range", "cashier"]},
            {"key": "lista_ventas", "label": "Lista de Ventas", "filters": ["range", "cashier", "payment_method"]},
            {"key": "ventas_diarias", "label": "Ventas Diarias", "filters": ["range", "cashier", "payment_method"]},
            {"key": "ventas_horas", "label": "Ventas por Horas", "filters": ["single_date", "cashier", "payment_method"]},
            {"key": "ventas_canceladas", "label": "Ventas Canceladas", "filters": ["range", "cashier"]},
            {"key": "devoluciones", "label": "Devoluciones", "filters": ["range", "cashier"]},
            {"key": "margen_beneficio", "label": "Margen de Beneficio", "filters": ["range", "cashier", "payment_method"]},
            {"key": "efectivo_inicio_caja", "label": "Efectivo en Inicio de Caja", "filters": ["range", "cashier"]},
            {"key": "movimientos_caja", "label": "Movimientos de Caja", "filters": ["range", "cashier"]},
            {"key": "descuentos_aplicados", "label": "Descuentos aplicados", "filters": ["range", "cashier", "payment_method"]},
        ]),
        ("Inventario", [
            {"key": "lista_productos", "label": "Lista de productos", "filters": []},
            {"key": "movimientos_inventario", "label": "Movimientos de inventarios", "filters": ["range"]},
            {"key": "stock_bajo", "label": "Stock Bajo", "filters": []},
        ]),
    ]

    informe_state = {"report_type": None, "report_def": None, "result": None, "filters_desc": ""}
    cashier_names: dict = {}

    informe_from_field, _ = _date_field(date.today().replace(day=1).isoformat(), label="Desde", width=140)
    informe_to_field, _   = _date_field(today_str, label="Hasta", width=140)
    informe_date_field, _ = _date_field(today_str, label="Fecha", width=140)
    informe_cashier_dropdown = ft.Dropdown(
        label="Cajero", width=190, border_color=PRIMARY, color=ft.colors.WHITE,
        bgcolor=BG_SURFACE, text_size=12,
        options=[ft.dropdown.Option("", "Todos")],
    )
    informe_payment_dropdown = ft.Dropdown(
        label="Forma de pago", width=170, border_color=PRIMARY, color=ft.colors.WHITE,
        bgcolor=BG_SURFACE, text_size=12,
        options=[ft.dropdown.Option(v, l) for v, l in PAYMENT_OPTIONS],
    )
    informe_payment_dropdown.value = ""
    informe_cashier_dropdown.value = ""

    informe_filter_area = ft.Row(spacing=10, wrap=True, vertical_alignment=ft.CrossAxisAlignment.END)
    informe_loading_ring = ft.ProgressRing(width=20, height=20, color=PRIMARY, visible=False)
    informe_results_view = ft.Column(spacing=8, scroll=ft.ScrollMode.AUTO, expand=True)
    informe_title_text = ft.Text("Seleccione un informe", size=15, color=ft.colors.WHITE,
                                  weight=ft.FontWeight.BOLD)

    def _load_cashiers():
        try:
            users = api.get_users()
            opts = [ft.dropdown.Option("", "Todos")]
            for u in users:
                name = u.get("full_name") or u.get("username", "")
                opts.append(ft.dropdown.Option(str(u["id"]), name))
                cashier_names[str(u["id"])] = name
            informe_cashier_dropdown.options = opts
        except APIError:
            pass

    def _fmt_date(iso_str: str) -> str:
        try:
            return date.fromisoformat(iso_str).strftime("%d/%m/%Y")
        except (ValueError, TypeError):
            return str(iso_str)

    def _build_filters_desc(item: dict, params: dict) -> str:
        """Describe en texto los filtros aplicados, para mostrarlos en la
        cabecera de los informes exportados."""
        filters = item.get("filters", [])
        parts = []
        if "range" in filters and "start" in params:
            parts.append(f"Periodo: {_fmt_date(params['start'])} - {_fmt_date(params['end'])}")
        if "single_date" in filters and "target_date" in params:
            parts.append(f"Fecha: {_fmt_date(params['target_date'])}")
        if "cashier" in filters:
            if "cashier_id" in params:
                name = cashier_names.get(str(params["cashier_id"]), f"#{params['cashier_id']}")
                parts.append(f"Cajero: {name}")
            else:
                parts.append("Cajero: Todos")
        if "payment_method" in filters:
            if params.get("payment_method"):
                label = PAYMENT_LABELS.get(params["payment_method"], params["payment_method"])
                parts.append(f"Forma de pago: {label}")
            else:
                parts.append("Forma de pago: Todas")
        return "   |   ".join(parts)

    def _format_informe_cell(value, col_type: str) -> str:
        if value is None or value == "":
            return "—"
        try:
            if col_type == "currency":
                return f"{currency}{float(value):,.2f}"
            if col_type == "percent":
                return f"{float(value):,.1f}%"
            if col_type == "qty":
                v = float(value)
                return f"{v:,.0f}" if v == int(v) else f"{v:,.2f}"
            if col_type == "number":
                return f"{int(value):,}"
            if col_type == "date":
                return str(value)[:10]
            if col_type == "datetime":
                return str(value).replace("T", " ")[:19]
        except (ValueError, TypeError):
            return str(value)
        return str(value)

    def _build_report_list():
        controls = []
        for group_label, items in REPORT_GROUPS:
            controls.append(ft.Container(
                padding=ft.padding.only(left=14, right=14, top=14, bottom=4),
                content=ft.Text(group_label.upper(), size=11, color=ft.colors.WHITE38,
                                 weight=ft.FontWeight.BOLD),
            ))
            for item in items:
                selected = informe_state["report_type"] == item["key"]
                controls.append(
                    ft.Container(
                        on_click=lambda e, it=item: _select_report(it),
                        padding=ft.padding.symmetric(horizontal=12, vertical=10),
                        margin=ft.margin.symmetric(horizontal=6),
                        border_radius=8,
                        bgcolor=(PRIMARY + "33") if selected else None,
                        ink=True,
                        content=ft.Text(
                            item["label"], size=13,
                            color=PRIMARY_LT if selected else ft.colors.WHITE70,
                            weight=ft.FontWeight.BOLD if selected else ft.FontWeight.NORMAL,
                        ),
                    )
                )
        return controls

    informe_list_view = ft.ListView(expand=True, spacing=2, controls=_build_report_list())

    def _build_filter_area():
        item = informe_state.get("report_def")
        informe_filter_area.controls.clear()
        if item:
            filters = item.get("filters", [])
            if "range" in filters:
                informe_filter_area.controls.append(informe_from_field)
                informe_filter_area.controls.append(informe_to_field)
            if "single_date" in filters:
                informe_filter_area.controls.append(informe_date_field)
            if "cashier" in filters:
                informe_filter_area.controls.append(informe_cashier_dropdown)
            if "payment_method" in filters:
                informe_filter_area.controls.append(informe_payment_dropdown)
        informe_filter_area.controls.append(
            ft.ElevatedButton("Mostrar informe", icon=ft.icons.VISIBILITY,
                              on_click=_generate_informe,
                              style=ft.ButtonStyle(bgcolor=PRIMARY, color=ft.colors.WHITE))
        )
        informe_filter_area.controls.append(informe_loading_ring)
        informe_filter_area.controls.append(
            ft.OutlinedButton("PDF", icon=ft.icons.PICTURE_AS_PDF, on_click=_export_informe_pdf)
        )
        informe_filter_area.controls.append(
            ft.OutlinedButton("Excel", icon=ft.icons.TABLE_CHART, on_click=_export_informe_excel)
        )

    def _select_report(item):
        informe_state["report_type"] = item["key"]
        informe_state["report_def"] = item
        informe_state["result"] = None
        informe_title_text.value = item["label"]
        informe_list_view.controls = _build_report_list()
        _build_filter_area()
        informe_results_view.controls = [
            _empty_placeholder("Configure los filtros y pulse 'Mostrar informe'")
        ]
        page.update()

    def _render_informe_result(result: dict):
        informe_results_view.controls.clear()
        columns = result.get("columns", [])
        rows = result.get("rows", [])
        if not rows:
            informe_results_view.controls.append(_empty_placeholder("No hay datos para los filtros seleccionados"))
            page.update()
            return

        dt_columns = [
            ft.DataColumn(
                ft.Text(c["label"], color=ft.colors.WHITE70, size=12),
                numeric=c["type"] in ("number", "qty", "currency", "percent"),
            )
            for c in columns
        ]
        dt_rows = []
        for r in rows:
            cells = [
                ft.DataCell(ft.Text(_format_informe_cell(r.get(c["key"]), c["type"]),
                                     size=12, color=ft.colors.WHITE70))
                for c in columns
            ]
            dt_rows.append(ft.DataRow(cells=cells))

        table = ft.DataTable(
            border=ft.border.all(1, ft.colors.WHITE12),
            border_radius=8,
            heading_row_color=BG_SURFACE,
            columns=dt_columns,
            rows=dt_rows,
        )
        informe_results_view.controls.append(ft.Row(scroll=ft.ScrollMode.AUTO, controls=[table]))
        informe_results_view.controls.append(
            ft.Text(f"{len(rows)} registro(s)", size=11, color=ft.colors.WHITE38)
        )
        page.update()

    def _generate_informe(e=None):
        item = informe_state.get("report_def")
        if not item:
            _show_snack("Seleccione un tipo de informe")
            return

        params = {}
        filters = item.get("filters", [])
        if "range" in filters:
            try:
                f = date.fromisoformat((informe_from_field.value or "").strip())
                t = date.fromisoformat((informe_to_field.value or "").strip())
            except ValueError:
                _show_snack("Formato de fecha inválido (use AAAA-MM-DD)")
                return
            if f > t:
                _show_snack("La fecha 'Desde' no puede ser posterior a 'Hasta'")
                return
            params["start"] = f.isoformat()
            params["end"] = t.isoformat()
        if "single_date" in filters:
            try:
                d = date.fromisoformat((informe_date_field.value or "").strip())
            except ValueError:
                _show_snack("Formato de fecha inválido (use AAAA-MM-DD)")
                return
            params["target_date"] = d.isoformat()
        if "cashier" in filters and informe_cashier_dropdown.value:
            params["cashier_id"] = int(informe_cashier_dropdown.value)
        if "payment_method" in filters and informe_payment_dropdown.value:
            params["payment_method"] = informe_payment_dropdown.value

        informe_loading_ring.visible = True
        page.update()
        try:
            result = api.get_custom_report(item["key"], **params)
            informe_state["result"] = result
            informe_state["filters_desc"] = _build_filters_desc(item, params)
            informe_title_text.value = result.get("title", item["label"])
            _render_informe_result(result)
        except APIError as ex:
            _show_snack(str(ex))
        finally:
            informe_loading_ring.visible = False
            page.update()

    # ── Exportación ────────────────────────────────────────────────────────

    def _open_exported_file(path: str):
        import subprocess, sys, os
        try:
            if sys.platform == "win32":
                os.startfile(path)
            elif sys.platform == "darwin":
                subprocess.run(["open", path], check=False)
            else:
                subprocess.run(["xdg-open", path], check=False)
        except Exception as ex:
            _show_snack(f"No se pudo abrir el archivo: {ex}")

    def _sanitize_pdf(text) -> str:
        return str(text).encode("latin-1", "replace").decode("latin-1")

    def _pdf_header(pdf, title: str, filters_desc: str = ""):
        """Dibuja la cabecera estándar de los PDF de reportes: datos de la
        tienda (a la izquierda) y el nombre del informe + fecha/hora de
        generación (a la derecha), seguidos de la descripción de los filtros
        aplicados (si los hay) y una línea separadora."""
        store_name    = cfg.get("store.name") or "Mi Tienda"
        store_address = cfg.get("store.address", "")
        store_phone   = cfg.get("store.phone", "")
        store_email   = cfg.get("store.email", "")
        store_tax_id  = cfg.get("store.tax_id", "")

        margin   = pdf.l_margin
        usable_w = pdf.w - 2 * margin
        half_w   = usable_w / 2
        line_h   = 5
        y0       = pdf.get_y()

        # ── Columna izquierda: datos de la tienda ──────────────────────────
        y = y0
        pdf.set_xy(margin, y)
        pdf.set_font("Helvetica", "B", 13)
        pdf.cell(half_w, 6, _sanitize_pdf(store_name))
        y += 6

        pdf.set_font("Helvetica", "", 8)
        contact = " · ".join(filter(None, [store_phone, store_email]))
        for line in filter(None, [store_address, contact,
                                   (f"RFC/NIT: {store_tax_id}" if store_tax_id else "")]):
            pdf.set_xy(margin, y)
            pdf.cell(half_w, line_h, _sanitize_pdf(line))
            y += line_h
        y_left_end = y

        # ── Columna derecha: nombre del informe y fecha de generación ──────
        x2 = margin + half_w
        y = y0
        pdf.set_xy(x2, y)
        pdf.set_font("Helvetica", "B", 13)
        pdf.cell(half_w, 6, _sanitize_pdf(title), align="R")
        y += 6

        pdf.set_font("Helvetica", "", 8)
        pdf.set_xy(x2, y)
        pdf.cell(half_w, line_h, "Generado: " + datetime.now().strftime("%d/%m/%Y %H:%M:%S"), align="R")
        y += line_h
        y_right_end = y

        # ── Filtros aplicados (línea completa, debajo de ambas columnas) ───
        y_cols_end = max(y_left_end, y_right_end)
        if filters_desc:
            pdf.set_xy(margin, y_cols_end)
            pdf.set_font("Helvetica", "I", 8)
            pdf.multi_cell(usable_w, line_h, _sanitize_pdf(f"Filtros: {filters_desc}"))
            y_cols_end = pdf.get_y()

        # ── Línea separadora ────────────────────────────────────────────────
        y_end = y_cols_end + 2
        pdf.set_draw_color(160, 160, 160)
        pdf.line(margin, y_end, pdf.w - margin, y_end)
        pdf.set_xy(margin, y_end + 4)

    def _export_informe_pdf(e=None):
        result = informe_state.get("result")
        if not result or not result.get("rows"):
            _show_snack("Genere un informe primero")
            return
        try:
            from fpdf import FPDF
        except ImportError:
            _show_snack("La librería fpdf2 no está instalada")
            return

        import tempfile, os
        columns = result["columns"]
        rows = result["rows"]
        landscape = len(columns) > 5

        pdf = FPDF(orientation="L" if landscape else "P", unit="mm", format="A4")
        pdf.add_page()
        _pdf_header(pdf, result.get("title", "Informe"), informe_state.get("filters_desc", ""))
        pdf.set_font("Helvetica", "", 9)

        page_width = pdf.w - 2 * pdf.l_margin

        # Ancho de columna proporcional al contenido: las columnas numéricas
        # (cantidades, montos, fechas) son angostas y las de texto (nombres,
        # categorías, etc.) obtienen el espacio restante según su contenido más
        # largo, para que los nombres de producto no se encimen con la siguiente
        # columna.
        NUMERIC_TYPES = ("currency", "qty", "number", "percent", "date")
        weights = []
        for c in columns:
            if c["type"] in NUMERIC_TYPES:
                weights.append(1.0)
            else:
                max_len = len(str(c["label"]))
                for r in rows:
                    val = _format_informe_cell(r.get(c["key"]), c["type"])
                    max_len = max(max_len, len(val))
                weights.append(max(1.5, min(max_len / 12, 4.0)))

        total_weight = sum(weights) or 1
        col_widths = [w / total_weight * page_width for w in weights]

        def fit_text(text: str, width: float) -> str:
            """Recorta el texto con '...' si no entra en el ancho de la celda."""
            if pdf.get_string_width(text) <= width - 2:
                return text
            while text and pdf.get_string_width(text + "...") > width - 2:
                text = text[:-1]
            return (text + "...") if text else "..."

        pdf.set_font("Helvetica", "B", 9)
        for c, w in zip(columns, col_widths):
            pdf.cell(w, 8, fit_text(_sanitize_pdf(c["label"]), w), border=1)
        pdf.ln()

        pdf.set_font("Helvetica", "", 8)
        for r in rows:
            for c, w in zip(columns, col_widths):
                val = _format_informe_cell(r.get(c["key"]), c["type"])
                pdf.cell(w, 7, fit_text(_sanitize_pdf(val), w), border=1)
            pdf.ln()

        tmp_path = os.path.join(tempfile.gettempdir(), f"informe_{result['report_type']}.pdf")
        pdf.output(tmp_path)
        _open_exported_file(tmp_path)
        _show_snack("PDF generado correctamente", SUCCESS)

    def _export_informe_excel(e=None):
        result = informe_state.get("result")
        if not result or not result.get("rows"):
            _show_snack("Genere un informe primero")
            return

        import tempfile, os
        columns = result["columns"]
        rows = result["rows"]
        base_path = os.path.join(tempfile.gettempdir(), f"informe_{result['report_type']}")

        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = (result.get("title", "Informe") or "Informe")[:31]
            ws.append([c["label"] for c in columns])
            for r in rows:
                ws.append([r.get(c["key"]) for c in columns])
            tmp_path = base_path + ".xlsx"
            wb.save(tmp_path)
        except ImportError:
            import csv
            tmp_path = base_path + ".csv"
            with open(tmp_path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow([c["label"] for c in columns])
                for r in rows:
                    writer.writerow([r.get(c["key"]) for c in columns])

        _open_exported_file(tmp_path)
        _show_snack("Archivo generado correctamente", SUCCESS)

    _build_filter_area()
    informe_results_view.controls.append(
        _empty_placeholder("Seleccione un tipo de informe a la izquierda")
    )

    informe_content = ft.Row(
        expand=True, spacing=12,
        vertical_alignment=ft.CrossAxisAlignment.STRETCH,
        controls=[
            ft.Container(
                width=300, bgcolor=BG_CARD, border_radius=12, padding=ft.padding.symmetric(vertical=4),
                content=informe_list_view,
            ),
            ft.Container(
                expand=True, bgcolor=BG_CARD, border_radius=12, padding=16,
                content=ft.Column(expand=True, spacing=10, controls=[
                    informe_title_text,
                    informe_filter_area,
                    ft.Divider(color=ft.colors.WHITE12),
                    informe_results_view,
                ]),
            ),
        ],
    )

    # ── Controles de filtros ──────────────────────────────────────────────────

    tab_index_ref = {"val": 0}

    def load_daily(e=None):
        d = (daily_date_field.value or today_str).strip()
        try:
            data = api.get_daily_report(d)
            render_daily(data)
        except APIError as ex:
            _show_snack(str(ex))

    def load_range(e=None):
        d_from = (date_from_field.value or today_str).strip()
        d_to   = (date_to_field.value or today_str).strip()
        try:
            data = api.get_range_report(d_from, d_to)
            render_range(data)
        except APIError as ex:
            _show_snack(str(ex))

    # Filas de filtros de fecha (Reporte Diario / Reporte por Rango)
    daily_filter_row = ft.Row(
        spacing=8,
        vertical_alignment=ft.CrossAxisAlignment.CENTER,
        controls=[
            ft.Text("Fecha:", color=ft.colors.WHITE54, size=12),
            daily_date_field,
            ft.ElevatedButton(
                "Buscar", icon=ft.icons.SEARCH,
                on_click=lambda e: load_daily(),
                style=ft.ButtonStyle(bgcolor=PRIMARY, color=ft.colors.WHITE),
            ),
        ],
    )
    range_filter_row = ft.Row(
        spacing=8,
        vertical_alignment=ft.CrossAxisAlignment.CENTER,
        controls=[
            ft.Text("Desde:", color=ft.colors.WHITE54, size=12),
            date_from_field,
            ft.Text("Hasta:", color=ft.colors.WHITE54, size=12),
            date_to_field,
            ft.ElevatedButton(
                "Buscar", icon=ft.icons.SEARCH,
                on_click=lambda e: load_range(),
                style=ft.ButtonStyle(bgcolor=PRIMARY, color=ft.colors.WHITE),
            ),
        ],
    )
    date_filter_row = ft.Row(spacing=8, vertical_alignment=ft.CrossAxisAlignment.CENTER,
                              controls=[daily_filter_row])

    daily_quick_row = ft.Row(
        spacing=4,
        vertical_alignment=ft.CrossAxisAlignment.CENTER,
        controls=[
            ft.Container(expand=True),
            ft.Text("Acceso rápido:", color=ft.colors.WHITE38, size=11),
            ft.TextButton("Hoy",  on_click=lambda _: _quick_date(0),
                          style=ft.ButtonStyle(color=PRIMARY_LT)),
            ft.TextButton("Ayer", on_click=lambda _: _quick_date(1),
                          style=ft.ButtonStyle(color=PRIMARY_LT)),
        ],
    )
    range_quick_row = ft.Row(
        spacing=4,
        vertical_alignment=ft.CrossAxisAlignment.CENTER,
        controls=[
            ft.Container(expand=True),
            ft.Text("Acceso rápido:", color=ft.colors.WHITE38, size=11),
            ft.TextButton("7 días",  on_click=lambda _: _quick_range(7),
                          style=ft.ButtonStyle(color=PRIMARY_LT)),
            ft.TextButton("30 días", on_click=lambda _: _quick_range(30),
                          style=ft.ButtonStyle(color=PRIMARY_LT)),
        ],
    )
    quick_access_row = ft.Row(spacing=4, vertical_alignment=ft.CrossAxisAlignment.CENTER,
                               controls=[daily_quick_row])

    date_filter_container  = ft.Container(
        bgcolor=BG_CARD,
        padding=ft.padding.only(left=16, right=16, top=10, bottom=6),
        content=ft.Row(spacing=8, vertical_alignment=ft.CrossAxisAlignment.CENTER, controls=[
            ft.Container(expand=True),
            date_filter_row,
        ]),
    )
    quick_access_container = ft.Container(
        bgcolor=BG_CARD,
        padding=ft.padding.only(left=16, right=16, bottom=8),
        content=quick_access_row,
    )

    def on_tab_change(e):
        idx = e.control.selected_index
        tab_index_ref["val"] = idx
        # Los filtros de fecha solo aplican a "Reporte Diario" y "Reporte por Rango"
        date_filter_container.visible  = idx in (1, 2)
        quick_access_container.visible = idx in (1, 2)
        if idx == 1:
            date_filter_row.controls = [daily_filter_row]
            quick_access_row.controls = [daily_quick_row]
        elif idx == 2:
            date_filter_row.controls = [range_filter_row]
            quick_access_row.controls = [range_quick_row]
        if idx == 0:
            load_monthly()
            load_period()
        elif idx == 1:
            daily_content.controls.clear()
            load_daily()
        elif idx == 2:
            report_content.controls.clear()
            load_range()
        page.update()

    def _quick_date(days_ago: int):
        d = (date.today() - timedelta(days=days_ago)).isoformat()
        daily_date_field.value = d
        page.update()
        load_daily()

    def _quick_range(days: int):
        d_to   = date.today().isoformat()
        d_from = (date.today() - timedelta(days=days - 1)).isoformat()
        date_from_field.value = d_from
        date_to_field.value = d_to
        page.update()
        load_range()

    # Carga inicial — pestaña "Resumen"
    date_filter_container.visible  = False
    quick_access_container.visible = False
    load_monthly()
    load_period()
    _load_cashiers()

    # Registrar los DatePicker en el overlay de la página
    page.overlay.extend(_date_pickers)

    return ft.Container(
        expand=True, bgcolor=BG_DARK,
        content=ft.Column(
            expand=True, spacing=0,
            controls=[
                # ── Fila 1: Título ──────────────────────────────────────────
                ft.Container(
                    bgcolor=BG_CARD,
                    padding=ft.padding.only(left=16, right=16, top=10, bottom=6),
                    content=ft.Row(
                        spacing=8,
                        vertical_alignment=ft.CrossAxisAlignment.CENTER,
                        controls=[
                            ft.Icon(ft.icons.ANALYTICS, color=PRIMARY),
                            ft.Text("Reportes Financieros", size=17,
                                    color=ft.colors.WHITE, weight=ft.FontWeight.BOLD),
                        ],
                    ),
                ),
                # ── Fila 2: Filtros de fecha (Diario / Rango) ────────────────
                date_filter_container,
                # ── Fila 3: Accesos rápidos (Diario / Rango) ─────────────────
                quick_access_container,
                ft.Divider(height=1, color=ft.colors.WHITE12),
                # ── Pestañas de reporte ───────────────────────────────────────
                ft.Tabs(
                    expand=True,
                    selected_index=0,
                    on_change=on_tab_change,
                    indicator_color=PRIMARY,
                    label_color=PRIMARY_LT,
                    unselected_label_color=ft.colors.WHITE54,
                    tabs=[
                        ft.Tab(
                            text="Resumen",
                            icon=ft.icons.DASHBOARD,
                            content=ft.Container(
                                expand=True,
                                padding=ft.padding.all(12),
                                content=ft.ListView(
                                    expand=True,
                                    controls=[dashboard_content],
                                    spacing=10,
                                ),
                            ),
                        ),
                        ft.Tab(
                            text="Reporte Diario",
                            icon=ft.icons.TODAY,
                            content=ft.Container(
                                expand=True,
                                padding=ft.padding.all(12),
                                content=ft.ListView(
                                    expand=True,
                                    controls=[daily_content],
                                    spacing=10,
                                ),
                            ),
                        ),
                        ft.Tab(
                            text="Reporte por Rango",
                            icon=ft.icons.DATE_RANGE,
                            content=ft.Container(
                                expand=True,
                                padding=ft.padding.all(12),
                                content=ft.ListView(
                                    expand=True,
                                    controls=[report_content],
                                    spacing=10,
                                ),
                            ),
                        ),
                        ft.Tab(
                            text="Informes",
                            icon=ft.icons.DESCRIPTION,
                            content=ft.Container(
                                expand=True,
                                padding=ft.padding.all(12),
                                content=informe_content,
                            ),
                        ),
                    ],
                ),
            ],
        ),
    )
